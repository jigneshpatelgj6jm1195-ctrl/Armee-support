"""
School Complaint Form - Data Exporter & Launcher
================================================
Run this script whenever the Master Excel sheet is updated.
Exports school data to school_data.json and starts a local HTTP server
so the form works in all browsers (fetch() requires http:// not file://).

Requirements:
    pip install openpyxl
"""

import json
import os
import sys
import threading
import webbrowser
import subprocess
import http.server
import socketserver
from pathlib import Path
from urllib.parse import urlparse

SCRIPT_DIR  = Path(__file__).parent
EXCEL_FILE  = SCRIPT_DIR / "School complaint format 23.6.25.xlsx"
JSON_FILE   = SCRIPT_DIR / "school_data.json"
MASTER_FILE = SCRIPT_DIR / "master_data.json"
HTML_FILE   = SCRIPT_DIR / "index.html"
PORT        = 8765


# ─────────────────────────────────────────────
# Custom HTTP handler: serves files + handles POSTs
# ─────────────────────────────────────────────
class FormHandler(http.server.SimpleHTTPRequestHandler):

    def log_message(self, format, *args):
        # Only log errors, not every GET request
        if args and str(args[1]) not in ('200', '304'):
            print(f"  [{args[1]}] {args[0]}")

    def do_POST(self):
        path = urlparse(self.path).path

        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
        except Exception as e:
            self._respond(400, {'error': f'Bad request: {e}'})
            return

        if path == '/upload_school_excel':
            self._handle_upload_school_excel(body)
            return

        try:
            data = json.loads(body)
        except Exception as e:
            self._respond(400, {'error': f'Bad request parsing JSON: {e}'})
            return

        if path == '/update_school':
            self._handle_update_school(data)
        elif path == '/update_master':
            self._handle_update_master(data)
        elif path == '/submit_complaint':
            self._handle_submit_complaint(data)
        elif path == '/update_complaints':
            self._handle_update_complaints(data)
        else:
            self._respond(404, {'error': 'Unknown endpoint'})

    def _handle_submit_complaint(self, data):
        """Append complaint payload to complaints.json file"""
        complaints_file = SCRIPT_DIR / "complaints.json"
        try:
            if complaints_file.exists():
                with open(complaints_file, 'r', encoding='utf-8') as f:
                    db = json.load(f)
            else:
                db = []
            
            if not isinstance(db, list):
                db = []

            db.append(data)
            with open(complaints_file, 'w', encoding='utf-8') as f:
                json.dump(db, f, ensure_ascii=False, indent=2)
            print(f"  [COMPLAINT] DISE {data.get('dise', '')} | School: {data.get('school', '')} | Equipment: {data.get('equipment', '')}")
            self._respond(200, {'ok': True})
        except Exception as e:
            self._respond(500, {'error': f'Could not save complaint: {e}'})

    def _handle_update_complaints(self, data):
        """Replace complaints.json with new content from admin panel"""
        complaints_file = SCRIPT_DIR / "complaints.json"
        try:
            with open(complaints_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print("  [COMPLAINTS] complaints.json updated")
            self._respond(200, {'ok': True})
        except Exception as e:
            self._respond(500, {'error': f'Could not write complaints.json: {e}'})

    def _handle_upload_school_excel(self, body):
        """Save uploaded school Excel sheet and rebuild school_data.json"""
        try:
            with open(EXCEL_FILE, 'wb') as f:
                f.write(body)
            print(f"  [EXCEL] Uploaded new master excel file. Parsing...")
            count = export_data()
            self._respond(200, {'ok': True, 'count': count})
        except Exception as e:
            self._respond(500, {'error': f'Failed to process excel: {e}'})

    def _handle_update_school(self, data):
        """Update a field (principal/mobile/district/block/pincode/address) in school_data.json"""
        dise      = str(data.get('dise', '')).strip()
        field     = str(data.get('field', '')).strip()
        new_value = str(data.get('newValue', '')).strip()
        old_value = str(data.get('oldValue', '')).strip()

        allowed_fields = {'principal', 'mobile', 'district', 'block', 'pincode', 'address'}
        if not dise or field not in allowed_fields or new_value is None:
            self._respond(400, {'error': 'Invalid parameters'})
            return

        try:
            with open(JSON_FILE, 'r', encoding='utf-8') as f:
                db = json.load(f)
        except Exception as e:
            self._respond(500, {'error': f'Could not read school_data.json: {e}'})
            return

        if dise not in db:
            self._respond(404, {'error': f'DISE {dise} not found'})
            return

        entries = db[dise]
        if not isinstance(entries, list):
            entries = [entries]
            db[dise] = entries

        updated = 0
        for record in entries:
            record[field] = new_value
            updated += 1

        try:
            with open(JSON_FILE, 'w', encoding='utf-8') as f:
                json.dump(db, f, ensure_ascii=False, separators=(',', ':'))
            print(f"  [UPDATE] DISE {dise}: {field} changed '{old_value}' -> '{new_value}'")
            self._respond(200, {'ok': True, 'updated': updated})
        except Exception as e:
            self._respond(500, {'error': f'Could not write school_data.json: {e}'})

    def _handle_update_master(self, data):
        """Replace master_data.json with new content from admin panel"""
        try:
            with open(MASTER_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print("  [MASTER] master_data.json updated")
            self._respond(200, {'ok': True})
        except Exception as e:
            self._respond(500, {'error': f'Could not write master_data.json: {e}'})

    def _respond(self, code, payload):
        body = json.dumps(payload).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', len(body))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()


# ─────────────────────────────────────────────
# EXPORT SCHOOL DATA
# ─────────────────────────────────────────────
def export_data():
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl

    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        input("Press Enter to exit...")
        sys.exit(1)

    print(f"Reading: {EXCEL_FILE.name}")
    wb = openpyxl.load_workbook(str(EXCEL_FILE), read_only=True, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip() == "Master Sheet":
            sheet_name = name
            break

    if not sheet_name:
        print(f"ERROR: 'Master Sheet' not found! Available: {wb.sheetnames}")
        input("Press Enter to exit...")
        sys.exit(1)

    ws = wb[sheet_name]
    data = {}
    count = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c is not None for c in row):
            continue
        if not row[2] or not row[1]:
            skipped += 1
            continue

        dise    = str(row[2]).strip()
        project = str(row[1]).strip()

        record = {
            "project":    project,
            "dise":       dise,
            "schoolCode": str(row[3]).strip() if row[3] else "",
            "district":   str(row[4]).strip() if row[4] else "",
            "block":      str(row[5]).strip() if row[5] else "",
            "school":     str(row[6]).strip() if row[6] else "",
            "principal":  str(row[7]).strip() if row[7] else "",
            "mobile":     str(row[8]).strip() if row[8] else "",
            "address":    str(row[9]).strip() if row[9] else "",
            "pincode":    str(row[10]).strip() if row[10] else "",
        }

        if dise not in data:
            data[dise] = []
        if not any(r["project"] == project for r in data[dise]):
            data[dise].append(record)
            count += 1

    wb.close()

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

    multi = sum(1 for v in data.values() if len(v) > 1)
    print(f"[OK] Exported {count} records ({len(data)} unique DISE codes) -> school_data.json")
    if skipped:
        print(f"     ({skipped} rows skipped)")
    if multi:
        print(f"     ({multi} DISE codes in multiple projects - all kept)")
    return count


# ─────────────────────────────────────────────
# LAUNCH LOCAL SERVER
# ─────────────────────────────────────────────
def launch_form():
    if not HTML_FILE.exists():
        print(f"ERROR: HTML form not found: {HTML_FILE}")
        return

    os.chdir(SCRIPT_DIR)

    global PORT
    server = None
    for try_port in range(PORT, PORT + 20):
        try:
            server = socketserver.TCPServer(("", try_port), FormHandler)
            server.allow_reuse_address = True
            PORT = try_port
            break
        except OSError:
            continue

    if not server:
        print("ERROR: Could not find a free port.")
        return

    url = f"http://localhost:{PORT}/index.html"
    admin_url = f"http://localhost:{PORT}/admin.html"
    print(f"\n[>>] Server started!")
    print(f"     Form:  {url}")
    print(f"     Admin: {admin_url}")
    print(f"\n     Keep this window open while using the form.")
    print(f"     Press Ctrl+C to stop.\n")

    threading.Timer(0.6, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[--] Server stopped.")
        server.shutdown()


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 55)
    print("  School Equipment Complaint Form")
    print("=" * 55)
    print()

    export_data()
    print()
    launch_form()
